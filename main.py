import requests
from openpyxl import load_workbook
import editpyxl
import json
import os
from time import time
from telegram.bot import Bot

# url/getUpdates get messages
# url/getMe bot information
# url/sendMessage?chat_id={chat_id}&text={myText} send message

class TelegramBot:
    def __init__(self):
        self.token = ''
        self.url = f'https://api.telegram.org/bot{self.token}/'
        self.user_states = []
        self.time = time()
        self.update_id = None

    def start(self):

        while True:
            messages = self.get_response(self.update_id)
            if messages:
                for message in messages:
                    
                    self.update_id = message['update_id']
                    chat_id = message['message']['from']['id']
                    user_text = message['message'].get('text', 'no text')
      
                    self.add_user(chat_id)
                    self.manage_message(chat_id,user_text)
         
    def get_response(self,update_id):
        request_link = f'{self.url}getUpdates?timeout=5'
        if update_id:
            request_link = f'{request_link}&offset={update_id + 1}'
        try:
            response = requests.get(request_link)
        except Exception as e:
            print(e)
  
        content = json.loads(response.content)['result']

        #remove messages before bot start
        for message in content:
            if message['message']['date'] < self.time:
                self.update_id = message['update_id']
                content.clear()
        
        return content

    def manage_message(self,chat_id,user_text):
        
        message_counter = self.get_counter(chat_id)
        
        if message_counter == 0:
            text = (f'Olá, eu sou o doguinho. O Assistente Virtual da Organnact.{os.linesep}'
            f'Para solicitar documentos, fale comigo, posso ter muitos documentos que podem te ajudar:{os.linesep}{os.linesep}'
            f'1 - Velocímetro{os.linesep}'
            '2 - Mapa da Mina')
            self.send_message(text,chat_id)

        elif message_counter == 1:
            text = 'Qual o distribuidor?'
            self.send_message(text,chat_id)

        elif message_counter == 2:
            self.scan_message(chat_id,user_text,'Distribuidor')
            text = 'Qual semana?'
            self.send_message(text,chat_id)

        elif message_counter == 3:
            self.scan_message(chat_id,user_text,'Semana')
            text = 'Qual a geração?'
            self.send_message(text,chat_id)

        elif message_counter == 4:
            self.scan_message(chat_id,user_text,'Geração')
            text = 'Qual a meta?:'
            self.send_message(text,chat_id)
        
        elif message_counter == 5:
            self.scan_message(chat_id,user_text,'Meta')
            text = 'Qual a positivação?'
            self.send_message(text,chat_id)
        
        elif message_counter == 6:
            self.scan_message(chat_id,user_text,'Positivação')
            self.excel_handler(chat_id)
            text = 'Aqui está seu documento'
            self.send_message(text,chat_id)
            self.send_file(chat_id,'VELOCIMETRO_EM_BRANCO_V2.xlsx')
            self.remove_user(chat_id)
    
    def scan_message(self,chat_id,user_text,name):
        for i,user_state in enumerate(self.user_states):
            if chat_id == self.user_states[i]['chatId']:
                self.user_states[i][name] = user_text
    
    def send_message(self,text,chat_id):
        send_link = f'{self.url}sendMessage?chat_id={chat_id}&text={text}'

        try:
            requests.post(send_link)
        except Exception as e:
            print(e)

    def send_file(self,chat_id,file_name):
        tbot = Bot(self.token)

        with open(file_name, "rb") as file:
          tbot.send_document(chat_id=chat_id, document=file, filename=file_name)
        if self.update_id:
            self.update_id += 1
        
    def get_counter(self,chat_id):
        for i,user_state in enumerate(self.user_states):
            if chat_id == self.user_states[i]['chatId']:
                return self.user_states[i]['messageCounter']
                    
    def excel_handler(self,chat_id):
        for i,user_state in enumerate(self.user_states):
            if chat_id == self.user_states[i]['chatId']:
                wb = editpyxl.Workbook()  
                wb.open('VELOCIMETRO_EM_BRANCO_V2.xlsx')
                ws=wb['RESUMO']

                ws.cell('B2').value= self.user_states[i]['Distribuidor']

                try:
                    ws.cell('D2').value= int(self.user_states[i]['Semana'])
                except:
                    print('Input is not number')
                else:
                    ws.cell('D2').value= self.user_states[i]['Semana']

                ws.cell('U2').value= self.user_states[i]['Geração']

                try:
                    ws.cell('F2').value= int(self.user_states[i]['Meta'])
                except:
                    print('Input is not number')
                else:
                    ws.cell('F2').value= self.user_states[i]['Meta']

                try:    
                    ws.cell('K2').value= int(self.user_states[i]['Positivação'])
                except:
                    print('Input is not number')
                else:
                    ws.cell('K2').value= self.user_states[i]['Positivação']

                wb.save(r'VELOCIMETRO_EM_BRANCO_V2.xlsx')
                wb.close()

    #add user or increase counter if user already exists
    def add_user(self,chat_id):

        user_information = {
            'chatId': chat_id,
            'messageCounter': 0,
            'Distribuidor': '',
            'Semana': '',
            'Geração': '',
            'Meta': 0,
            'Positivação': 0,
        }

        #Add new user or increase counter
        last_index = len(self.user_states) - 1
        for i,user_state in enumerate(self.user_states):
            if(self.user_states[i]['chatId'] == chat_id):
                self.user_states[i]['messageCounter'] += 1
                break
            elif(i == last_index):
                self.user_states.append(user_information)
                break
        
        #List empty special case
        if(not self.user_states):
            self.user_states.append(user_information)
       
    def remove_user(self,chat_id):
        for i,user_state in enumerate(self.user_states):
            if chat_id == self.user_states[i]['chatId']:
                del self.user_states[i]

bot = TelegramBot()
bot.start()